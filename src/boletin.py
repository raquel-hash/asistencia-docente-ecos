import tkinter as tk
from tkinter import messagebox
import openpyxl

def generar_boletin():
    ci = entry_ci.get()
    fecha = entry_fecha.get()
    horas_trabajadas = int(entry_horas.get())
    retrasos = entry_retrasos.get()
    
    # Buscar docente
    wb_docentes = openpyxl.load_workbook('../data/docentes.xlsx')
    ws_docentes = wb_docentes.active
    docente_encontrado = False
    for row in ws_docentes.iter_rows(min_row=2, values_only=True):
        if row[1] == ci:
            docente_encontrado = True
            nombre = row[0]
            pago_por_hora = float(row[2])
            break
    
    if not docente_encontrado:
        messagebox.showwarning("Error", "Docente no encontrado.")
        return
    
    # Cálculos de pago
    descuento_por_retraso = 5.00
    total_descuentos = retrasos.count('00:05:00') * descuento_por_retraso
    total_ganado = horas_trabajadas * pago_por_hora - total_descuentos
    
    # Guardar en Excel
    wb_pagos = openpyxl.load_workbook('../data/pagos.xlsx')
    ws_pagos = wb_pagos.active
    ws_pagos.append([nombre, ci, fecha, horas_trabajadas, retrasos, total_descuentos, total_ganado])
    wb_pagos.save('../data/pagos.xlsx')
    
    messagebox.showinfo("Éxito", "Boletín generado y guardado en Excel.")

def abrir_generador_boletines():
    root = tk.Toplevel()
    root.title("Generador de Boletines de Pago")

    # Estilos
    root.geometry("400x200")
    root.configure(bg='#f0f0f0')

    tk.Label(root, text="C.I.", bg='#f0f0f0').grid(row=0, column=0, padx=10, pady=5)
    global entry_ci
    entry_ci = tk.Entry(root)
    entry_ci.grid(row=0, column=1, padx=10, pady=5)

    tk.Label(root, text="Fecha (YYYY-MM-DD)", bg='#f0f0f0').grid(row=1, column=0, padx=10, pady=5)
    global entry_fecha
    entry_fecha = tk.Entry(root)
    entry_fecha.grid(row=1, column=1, padx=10, pady=5)

    tk.Label(root, text="Horas Trabajadas", bg='#f0f0f0').grid(row=2, column=0, padx=10, pady=5)
    global entry_horas
    entry_horas = tk.Entry(root)
    entry_horas.grid(row=2, column=1, padx=10, pady=5)

    tk.Label(root, text="Retrasos (HH:MM:SS)", bg='#f0f0f0').grid(row=3, column=0, padx=10, pady=5)
    global entry_retrasos
    entry_retrasos = tk.Entry(root)
    entry_retrasos.grid(row=3, column=1, padx=10, pady=5)

    tk.Button(root, text="Generar Boletin", command=generar_boletin, bg='#4CAF50', fg='white').grid(row=4, column=0, columnspan=2, pady=10)
