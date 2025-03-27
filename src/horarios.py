import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os
from datetime import datetime

archivo_materias = os.path.join(os.path.dirname(__file__), '../data/materias.xlsx')
archivo_horarios = os.path.join(os.path.dirname(__file__), '../data/horarios.xlsx')
archivo_docentes = os.path.join(os.path.dirname(__file__), '../data/docentes.xlsx')

def verificar_archivo_horarios():
    """Verifica si el archivo de horarios existe, si no lo crea vacío con encabezados."""
    if not os.path.exists(archivo_horarios):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["C.I.", "Nombre", "Materia", "Día", "Hora Inicio", "Hora Fin", "Horas Trabajadas"])
        wb.save(archivo_horarios)

def actualizar_lista_horarios(tree, filtro_ci=None):
    """Actualiza la tabla de horarios en la ventana."""
    for row in tree.get_children():
        tree.delete(row)

    wb = openpyxl.load_workbook(archivo_horarios)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if filtro_ci and row[0] != filtro_ci:
            continue
        tree.insert("", "end", values=row)

def calcular_horas_trabajadas(hora_inicio, hora_fin):
    """Calcula las horas trabajadas dada la hora de inicio y la hora de fin."""
    formato = "%H:%M"
    h_inicio = datetime.strptime(hora_inicio, formato)
    h_fin = datetime.strptime(hora_fin, formato)
    diferencia = h_fin - h_inicio
    horas_trabajadas = diferencia.total_seconds() / 3600
    return horas_trabajadas

def validar_hora(hora):
    """Valida que la hora esté en el formato HH:MM."""
    try:
        datetime.strptime(hora, "%H:%M")
        return True
    except ValueError:
        return False

def abrir_gestion_horarios(root_menu):
    """Abre la ventana para gestionar los horarios de los docentes."""
    verificar_archivo_horarios()

    # Evitar múltiples ventanas abiertas
    if hasattr(abrir_gestion_horarios, "ventana") and abrir_gestion_horarios.ventana.winfo_exists():
        abrir_gestion_horarios.ventana.lift()
        return

    root_menu.withdraw()
    root = tk.Toplevel()
    abrir_gestion_horarios.ventana = root
    root.title("Gestión de Horarios")
    root.geometry("1000x600")
    root.state('zoomed')
    root.configure(bg='#e0e0e0')
    
    style = ttk.Style()
    style.configure("Treeview.Heading", font=('Arial', 12, 'bold'), background='#d9d9d9', foreground='black')
    style.configure("Treeview", font=('Arial', 10), rowheight=25, background='#f0f0f0', foreground='black', fieldbackground='#f0f0f0')

    # Frame principal
    frame_main = tk.Frame(root, bg='#e0e0e0')
    frame_main.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

    frame_form = tk.Frame(frame_main, bg='#e0e0e0')
    frame_form.pack(side=tk.LEFT, fill=tk.Y, padx=20, pady=20)

    frame_table = tk.Frame(frame_main, bg='#e0e0e0')
    frame_table.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=20, pady=20)

    # Título
    # tk.Label(frame_main, text="Gestión de Horarios", font=("Arial", 24), bg='#e0e0e0').pack(pady=20)

    # Formulario de registro de horarios
    labels = ["C.I. y Nombre", "Materia", "Día", "Hora Inicio (HH:MM)", "Hora Fin (HH:MM)"]
    entries = []

    # Obtener lista de docentes y materias
    wb_docentes = openpyxl.load_workbook(archivo_docentes)
    ws_docentes = wb_docentes.active
    docentes = [(row[0], row[1]) for row in ws_docentes.iter_rows(min_row=2, values_only=True)]
    docentes_combo = [f"{docente[0]} - {docente[1]}" for docente in docentes]

    wb_materias = openpyxl.load_workbook(archivo_materias)
    ws_materias = wb_materias.active
    materias = [row[0] for row in ws_materias.iter_rows(min_row=2, values_only=True)]

    # Días disponibles
    dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]

    for i, label in enumerate(labels):
        tk.Label(frame_form, text=label + ":", bg='#e0e0e0', font=('Arial', 10)).grid(row=i, column=0, padx=5, pady=5, sticky="e")
        
        if "C.I. y Nombre" in label:
            combo_ci_nombre = ttk.Combobox(frame_form, values=docentes_combo, width=30)
            entry = combo_ci_nombre
        elif "Materia" in label:
            entry = ttk.Combobox(frame_form, values=materias, width=30)
        elif "Día" in label:
            entry = ttk.Combobox(frame_form, values=dias, width=30)
        else:
            entry = tk.Entry(frame_form, width=30)
        
        entry.grid(row=i, column=1, padx=5, pady=5)
        entries.append(entry)

    def agregar_horario():
        """Agrega un nuevo horario al archivo y actualiza la lista."""
        datos = [entry.get() for entry in entries]

        if not all(datos):
            messagebox.showwarning("Campos incompletos", "Todos los campos son obligatorios.")
            return

        ci = datos[0].split(" - ")[0]  # Extraer solo el C.I.
        nombre = datos[0].split(" - ")[1]  # Extraer el nombre
        datos[0] = ci
        datos.insert(1, nombre)  # Insertar el nombre en la posición correcta

        hora_inicio = datos[4]
        hora_fin = datos[5]

        if not validar_hora(hora_inicio) or not validar_hora(hora_fin):
            messagebox.showerror("Formato incorrecto", "La hora debe estar en formato HH:MM")
            return

        horas_trabajadas = calcular_horas_trabajadas(hora_inicio, hora_fin)
        datos.append(str(horas_trabajadas))

        wb = openpyxl.load_workbook(archivo_horarios)
        ws = wb.active
        ws.append(datos)
        wb.save(archivo_horarios)

        actualizar_lista_horarios(tree)

        for entry in entries:
            entry.delete(0, tk.END)

    def eliminar_horario():
        """Elimina el horario seleccionado de la tabla y del archivo."""
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("Seleccionar horario", "Debe seleccionar un horario para eliminar.")
            return
        
        horario = tree.item(selected_item, "values")

        wb = openpyxl.load_workbook(archivo_horarios)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if all(row[i].value == horario[i] for i in range(len(horario))):
                ws.delete_rows(row[0].row)
                break
        wb.save(archivo_horarios)

        actualizar_lista_horarios(tree)

    def filtrar_horarios():
        """Filtra los horarios por C.I. del docente seleccionado."""
        seleccion = filtro_ci_nombre.get()
        ci = seleccion.split(" - ")[0] if seleccion else None
        actualizar_lista_horarios(tree, filtro_ci=ci)

    tk.Button(frame_form, text="Agregar Horario", command=agregar_horario, bg='#212121', fg='white').grid(row=len(labels)+1, columnspan=2, pady=10)
    tk.Button(frame_form, text="Eliminar Horario", command=eliminar_horario, bg='#212121', fg='white').grid(row=len(labels)+2, columnspan=2, pady=10)

    # Filtro de horarios
    filtro_frame = tk.Frame(frame_table, bg='#e0e0e0')
    filtro_frame.pack(pady=5, anchor='w')

    tk.Label(filtro_frame, text="Filtrar por C.I.:", bg='#e0e0e0', font=('Arial', 10)).pack(side=tk.LEFT, padx=5)
    filtro_ci_nombre = ttk.Combobox(filtro_frame, values=docentes_combo, width=30)
    filtro_ci_nombre.pack(side=tk.LEFT, padx=5)
    tk.Button(filtro_frame, text="Filtrar", command=filtrar_horarios, bg='#212121', fg='white').pack(side=tk.LEFT, padx=5)

    # Tabla de horarios
    columns = ["C.I.", "Nombre", "Materia", "Día", "Hora Inicio", "Hora Fin", "Horas Trabajadas"]
    tree = ttk.Treeview(frame_table, columns=columns, show="headings", height=10)
    
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=110, anchor="center")
    
    tree.pack(pady=10, fill=tk.BOTH, expand=True)

    actualizar_lista_horarios(tree)

    # Botón para volver al menú
    tk.Button(root, text="Volver al Menú", command=lambda: volver_al_menu(root, root_menu), 
              bg='#212121', fg='white', width=20).pack(pady=10)

    root.protocol("WM_DELETE_WINDOW", lambda: volver_al_menu(root, root_menu))  # Manejar cierre de ventana
    root.mainloop()

def volver_al_menu(root_actual, root_menu):
    """Cierra la ventana actual y muestra nuevamente el menú principal."""
    root_actual.destroy()
    root_menu.deiconify()
    root_menu.state('zoomed')  # Maximiza la ventana del menú principal al volver
