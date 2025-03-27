import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os
from datetime import datetime

archivo_asistencia = os.path.join(os.path.dirname(__file__), '../data/asistencia.xlsx')
archivo_docentes = os.path.join(os.path.dirname(__file__), '../data/docentes.xlsx')

def verificar_archivo_asistencia():
    """Verifica si el archivo de asistencia existe, si no lo crea vacío con encabezados."""
    if not os.path.exists(archivo_asistencia):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["C.I.", "Nombre", "Fecha", "Hora Entrada", "Hora Salida"])
        wb.save(archivo_asistencia)
        print("Archivo de asistencia creado con encabezados.")

def actualizar_lista_asistencia(tree):
    """Actualiza la tabla de asistencia en la ventana."""
    for row in tree.get_children():
        tree.delete(row)

    wb = openpyxl.load_workbook(archivo_asistencia)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = list(row)
        if values[4] is None or values[4] == "":
            values[4] = "En Proceso"
            tag = 'in_progress'
        else:
            tag = ''

        if isinstance(values[2], datetime):
            values[2] = values[2].strftime("%Y-%m-%d")
        if isinstance(values[3], datetime):
            values[3] = values[3].strftime("%H:%M:%S")
        if isinstance(values[4], datetime):
            values[4] = values[4].strftime("%H:%M:%S")

        tree.insert("", "end", values=values, tags=(tag,))
    print("Lista de asistencia actualizada en la tabla.", values)

def registrar_entrada(ci, nombre):
    """Registra la hora de entrada del docente."""
    fecha = datetime.now().strftime("%Y-%m-%d")
    hora_entrada = datetime.now().strftime("%H:%M:%S")
    wb = openpyxl.load_workbook(archivo_asistencia)
    ws = wb.active
    ws.append([ci, nombre, fecha, hora_entrada, None])
    wb.save(archivo_asistencia)
    print(f"Hora de entrada registrada para {ci} - {nombre}")

def registrar_salida(ci):
    """Registra la hora de salida del docente."""
    wb = openpyxl.load_workbook(archivo_asistencia)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == ci and (row[4].value is None or row[4].value == ""):
            row[4].value = datetime.now().strftime("%H:%M:%S")
            break
    wb.save(archivo_asistencia)
    print(f"Hora de salida registrada para {ci}")

def abrir_registro_asistencia(root_menu):
    """Abre la ventana para registrar la asistencia de los docentes."""
    verificar_archivo_asistencia()

    # Evitar múltiples ventanas abiertas
    if hasattr(abrir_registro_asistencia, "ventana") and abrir_registro_asistencia.ventana.winfo_exists():
        abrir_registro_asistencia.ventana.lift()
        return

    root_menu.withdraw()
    root = tk.Toplevel()
    abrir_registro_asistencia.ventana = root
    root.title("Registro de Asistencia")
    root.geometry("1000x600")
    root.state('zoomed')
    root.configure(bg='#e0e0e0')
    
    style = ttk.Style()
    style.configure("Treeview.Heading", font=('Arial', 12, 'bold'), background='#d9d9d9', foreground='black')
    style.configure("Treeview", font=('Arial', 10), rowheight=25, background='#f0f0f0', foreground='black', fieldbackground='#f0f0f0')
    style.map('Treeview', background=[('selected', '#347083')], foreground=[('selected', 'white')])

    # Frame principal
    frame_main = tk.Frame(root, bg='#e0e0e0')
    frame_main.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

    frame_table = tk.Frame(frame_main, bg='#e0e0e0')
    frame_table.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

    # Tabla de asistencia
    columns = ["C.I.", "Nombre", "Fecha", "Hora Entrada", "Hora Salida"]
    tree = ttk.Treeview(frame_table, columns=columns, show="headings", height=10)
    
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=110, anchor="center")
    
    tree.tag_configure('in_progress', background='lightgoldenrodyellow')
    tree.pack(pady=10, fill=tk.BOTH, expand=True)

    actualizar_lista_asistencia(tree)

    # Botón para volver al menú
    tk.Button(root, text="Volver al Menú", command=lambda: volver_al_menu(root, root_menu), 
              bg='#212121', fg='white', width=20).pack(pady=10)

    root.protocol("WM_DELETE_WINDOW", lambda: volver_al_menu(root, root_menu))  # Manejar cierre de ventana

    def mostrar_modal():
        """Muestra el modal para registrar la asistencia del docente."""
        modal = tk.Toplevel(root)
        modal.title("Registrar Asistencia")
        # modal.geometry("400x200")  # Tamaño inicial de la ventana modal
        modal.configure(bg='#e0e0e0')
        modal.transient(root)
        modal.grab_set()
        modal.focus_set()

        # Frame para centrar los elementos
        frame_modal = tk.Frame(modal, bg='#e0e0e0')
        frame_modal.pack(expand=True, fill=tk.BOTH, padx=20, pady=20)

        tk.Label(frame_modal, text="C.I.:", bg='#e0e0e0', font=('Arial', 12)).pack(pady=5)
        entry_ci = tk.Entry(frame_modal, width=30, justify='center', font=('Arial', 12))
        entry_ci.pack(pady=5)
        entry_ci.focus_set()

        def registrar_asistencia(event=None):
            """Registra la asistencia del docente."""
            ci = entry_ci.get().strip()

            if not ci:
                messagebox.showwarning("Campo incompleto", "El campo C.I. es obligatorio.", parent=modal)
                return

            wb_docentes = openpyxl.load_workbook(archivo_docentes)
            ws_docentes = wb_docentes.active
            docentes_dict = {str(row[0]).strip(): str(row[1]).strip() for row in ws_docentes.iter_rows(min_row=2, values_only=True)}

            if ci not in docentes_dict:
                messagebox.showerror("C.I. no encontrado", "El C.I. ingresado no corresponde a ningún docente.", parent=modal)
                return

            nombre = docentes_dict[ci]
            wb = openpyxl.load_workbook(archivo_asistencia)
            ws = wb.active

            # Verificar si ya se registró la entrada hoy sin salida
            entrada_registrada = False
            for row in ws.iter_rows(min_row=2):
                if row[0].value == ci and row[2].value == datetime.now().strftime("%Y-%m-%d") and (row[4].value is None or row[4].value == ""):
                    entrada_registrada = True
                    row[4].value = datetime.now().strftime("%H:%M:%S")
                    print(f"Hora de salida registrada para {ci} - {nombre}")
                    break

            if not entrada_registrada:
                registrar_entrada(ci, nombre)
                messagebox.showinfo("Éxito", "Hora de entrada registrada correctamente.", parent=modal)
            else:
                wb.save(archivo_asistencia)
                messagebox.showinfo("Éxito", "Hora de salida registrada correctamente.", parent=modal)

            actualizar_lista_asistencia(tree)

            entry_ci.delete(0, tk.END)

        tk.Button(frame_modal, text="Registrar Asistencia", command=registrar_asistencia, bg='#212121', fg='white', font=('Arial', 12)).pack(pady=10)

        # Bind the Enter key to the registrar_asistencia function
        entry_ci.bind('<Return>', registrar_asistencia)

        # Actualizar la ventana modal para centrarla después de empaquetar los elementos
        modal.update_idletasks()
        width = modal.winfo_width()
        height = modal.winfo_height()
        x = (modal.winfo_screenwidth() // 2) - (width // 2)
        y = (modal.winfo_screenheight() // 2) - (height // 2)
        modal.geometry(f'{width}x{height}+{x}+{y}')

    root.after(500, mostrar_modal)
    root.mainloop()

def volver_al_menu(root_actual, root_menu):
    """Cierra la ventana actual y muestra nuevamente el menú principal."""
    root_actual.destroy()
    root_menu.deiconify()
    root_menu.state('zoomed')  # Maximiza la ventana del menú principal al volver
