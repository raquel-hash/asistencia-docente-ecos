import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os
from datetime import datetime

archivo_docentes = os.path.join(os.path.dirname(__file__), '../data/docentes.xlsx')

def verificar_archivo():
    """Verifica si el archivo de docentes existe, si no lo crea vacío con encabezados."""
    if not os.path.exists(archivo_docentes):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["C.I.", "Nombre", "Especialidad", "Pago por Hora", "Celular"])
        wb.save(archivo_docentes)

def actualizar_lista(tree):
    """Actualiza la tabla de docentes en la ventana."""
    for row in tree.get_children():
        tree.delete(row)

    wb = openpyxl.load_workbook(archivo_docentes)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        tree.insert("", "end", values=row)

def abrir_lista_docentes(root_menu):
    """Abre la ventana con la lista de docentes y el formulario de registro."""
    verificar_archivo()

    if hasattr(abrir_lista_docentes, "ventana") and abrir_lista_docentes.ventana.winfo_exists():
        abrir_lista_docentes.ventana.lift()
        return

    root_menu.withdraw()
    root = tk.Toplevel()
    abrir_lista_docentes.ventana = root
    root.title("Lista de Docentes")
    root.geometry("1000x600")
    root.state('zoomed')
    root.configure(bg='#e0e0e0')
    
    style = ttk.Style()
    style.configure("Treeview.Heading", font=('Arial', 12, 'bold'), background='#d9d9d9', foreground='black')
    style.configure("Treeview", font=('Arial', 10), rowheight=25, background='#f0f0f0', foreground='black', fieldbackground='#f0f0f0')

    # Frame principal
    frame_main = tk.Frame(root, bg='#e0e0e0')
    frame_main.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

    frame_form = tk.Frame(frame_main, bg='#e0e0e0')
    frame_form.pack(side=tk.LEFT, fill=tk.Y, padx=20, pady=20)

    frame_table = tk.Frame(frame_main, bg='#e0e0e0')
    frame_table.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=20, pady=20)

    # Título
    tk.Label(frame_main, text="Gestión de Docentes", font=("Arial", 24), bg='#e0e0e0').pack(pady=20)

    # Formulario de registro de docentes
    labels = ["C.I.", "Nombre", "Especialidad", "Pago por Hora", "Celular"]
    entries = []

    for i, label in enumerate(labels):
        tk.Label(frame_form, text=label + ":", bg='#e0e0e0', font=('Arial', 10)).grid(row=i, column=0, padx=5, pady=5, sticky="e")
        entry = tk.Entry(frame_form)
        entry.grid(row=i, column=1, padx=5, pady=5)
        entries.append(entry)

    def agregar_docente():
        """Agrega un nuevo docente al archivo y actualiza la lista."""
        datos = [entry.get() for entry in entries]

        if not all(datos):
            messagebox.showwarning("Campos incompletos", "Todos los campos son obligatorios.")
            return

        wb = openpyxl.load_workbook(archivo_docentes)
        ws = wb.active
        ws.append(datos)
        wb.save(archivo_docentes)

        actualizar_lista(tree)

        for entry in entries:
            entry.delete(0, tk.END)

    def editar_docente():
        """Permite editar los datos de un docente seleccionado."""
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("Seleccionar docente", "Debe seleccionar un docente para editar.")
            return

        values = tree.item(selected_item, "values")
        for entry, value in zip(entries, values):
            entry.delete(0, tk.END)
            entry.insert(0, value)

        tk.Button(frame_form, text="Guardar Cambios", command=guardar_cambios, bg='#212121', fg='white').grid(row=len(labels), columnspan=2, pady=10)

    def guardar_cambios():
        """Guarda los cambios realizados a un docente existente."""
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("Seleccionar docente", "Debe seleccionar un docente para guardar los cambios.")
            return

        datos = [entry.get() for entry in entries]

        if not all(datos):
            messagebox.showwarning("Campos incompletos", "Todos los campos son obligatorios.")
            return

        wb = openpyxl.load_workbook(archivo_docentes)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            if row[0].value == datos[0]:
                for i, cell in enumerate(row):
                    if i < len(datos):
                        cell.value = datos[i]
                break

        wb.save(archivo_docentes)
        actualizar_lista(tree)

        for entry in entries:
            entry.delete(0, tk.END)

    tk.Button(frame_form, text="Agregar Docente", command=agregar_docente, bg='#212121', fg='white').grid(row=len(labels)+1, columnspan=2, pady=10)
    tk.Button(frame_form, text="Editar Docente", command=editar_docente, bg='#212121', fg='white').grid(row=len(labels)+2, columnspan=2, pady=10)

    # Tabla de docentes
    columns = ["C.I.", "Nombre", "Especialidad", "Pago por Hora", "Celular"]
    tree = ttk.Treeview(frame_table, columns=columns, show="headings", height=10)
    
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=110, anchor="center")
    
    tree.pack(pady=10, fill=tk.BOTH, expand=True)

    actualizar_lista(tree)

    # Botón para volver al menú
    tk.Button(root, text="Volver al Menú", command=lambda: volver_al_menu(root, root_menu), 
              bg='#212121', fg='white', width=20).pack(pady=10)

    root.protocol("WM_DELETE_WINDOW", lambda: volver_al_menu(root, root_menu))  # Manejar cierre de ventana
    root.mainloop()

def volver_al_menu(root_actual, root_menu):
    """Cierra la ventana actual y muestra nuevamente el menú principal."""
    root_actual.destroy()
    root_menu.deiconify()
    root_menu.state('zoomed')  # Maximiza la ventana del menú principal al volver
