import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os
from datetime import datetime, timedelta, time

archivo_asistencia = os.path.join(os.path.dirname(__file__), '../data/asistencia.xlsx')
archivo_docentes = os.path.join(os.path.dirname(__file__), '../data/docentes.xlsx')
archivo_horarios = os.path.join(os.path.dirname(__file__), '../data/horarios.xlsx')
plantilla_excel = os.path.join(os.path.dirname(__file__), '../plantilla/report.xlsx')

def verificar_archivo_asistencia():
    """Verifica si el archivo de asistencia existe, si no lo crea vacío con encabezados."""
    if not os.path.exists(archivo_asistencia):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["C.I.", "Nombre", "Fecha", "Hora Entrada", "Hora Salida"])
        wb.save(archivo_asistencia)

def obtener_docentes():
    """Obtiene la lista de docentes del archivo de docentes."""
    wb_docentes = openpyxl.load_workbook(archivo_docentes)
    ws_docentes = wb_docentes.active
    docentes = {}
    for row in ws_docentes.iter_rows(min_row=2, values_only=True):
        try:
            ci = str(row[0]).strip()
            nombre = str(row[1]).strip()
            pago_por_hora = float(row[3]) if row[3] is not None else 0.0  # Índice 3 para la columna "Pago por Hora"
            docentes[ci] = (nombre, pago_por_hora)
        except ValueError as e:
            print(f"Error al procesar la fila {row}: {e}")
    return docentes

def obtener_horario(ci):
    """Obtiene el horario del docente del archivo de horarios."""
    horario = {}
    if os.path.exists(archivo_horarios):
        wb_horarios = openpyxl.load_workbook(archivo_horarios)
        ws_horarios = wb_horarios.active
        for row in ws_horarios.iter_rows(min_row=2, values_only=True):
            if str(row[0]).strip() == ci:
                materia = row[2]
                dia = row[3].lower()  # Asegurarse de que el día esté en minúsculas
                hora_inicio = row[4]
                hora_fin = row[5]
                if dia not in horario:
                    horario[dia] = []
                horario[dia].append((materia, hora_inicio, hora_fin))
    return horario

def obtener_dias_mes(year, month, dia):
    """Obtiene todos los días específicos (e.g., lunes) de un mes y año dados."""
    dias = []
    date = datetime(year, month, 1)
    target_day = dia.lower()

    # Mapping of days to numbers
    days_map = {
        'lunes': 0,
        'martes': 1,
        'miércoles': 2,
        'jueves': 3,
        'viernes': 4,
        'sábado': 5,
        'domingo': 6
    }

    # Get the target day number
    target_day_num = days_map[target_day]

    # Find the first occurrence of the target day in the month
    while date.weekday() != target_day_num:
        date += timedelta(days=1)

    # Collect all occurrences of the target day in the month
    while date.month == month:
        dias.append(date.strftime("%Y-%m-%d"))
        date += timedelta(days=7)

    return dias

def obtener_anios_disponibles():
    """Obtiene los años disponibles en el archivo de asistencia."""
    if not os.path.exists(archivo_asistencia):
        return []

    wb = openpyxl.load_workbook(archivo_asistencia)
    ws = wb.active
    anios = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            fecha = row[2]
            if isinstance(fecha, str):
                fecha = datetime.strptime(fecha, "%Y-%m-%d")
            anios.add(fecha.year)
        except Exception as e:
            print(f"Error procesando la fecha {row[2]}: {e}")

    return sorted(anios)

def calcular_retraso(entrada, hora_programada):
    if isinstance(entrada, str):
        entrada = datetime.strptime(entrada, "%H:%M:%S")

    if isinstance(hora_programada, str):
        hora_programada = datetime.strptime(hora_programada, "%H:%M:%S")

    if entrada.time() > hora_programada.time():
        retraso = (entrada - hora_programada).seconds // 60
    else:
        retraso = 0
    return retraso

def calcular_deduccion(retraso_minutos):
    """Calcula la deducción en función de los minutos de retraso."""
    if retraso_minutos <= 5:
        return 5
    else:
        return 10

def formatear_retraso(retraso_minutos):
    """Formatea el retraso en formato HH:MM:SS."""
    horas, minutos = divmod(retraso_minutos, 60)
    return f"{horas:02}:{minutos:02}:00"

def generar_reporte(ci, mes, year):
    """Genera el reporte de horas trabajadas y deducciones para un docente y mes específico."""
    wb = openpyxl.load_workbook(archivo_asistencia)
    ws = wb.active
    
    total_horas = 0
    deducciones = 0
    registros = []
    total_retrasos = timedelta()

    # Obtener el horario del docente
    horario = obtener_horario(ci)

    # Obtener todos los días específicos del mes según el horario del docente
    for dia, materias in horario.items():
        for materia, hora_inicio, hora_fin in materias:
            dia_lower = dia.lower()
            dias_mes = obtener_dias_mes(year, mes, dia_lower)
            for fecha in dias_mes:
                registros.append([fecha, dia_lower, materia, 0, "00:00:00", 0, "PRESENCIAL"])

    # Actualizar los registros con los datos de asistencia
    for row in ws.iter_rows(min_row=2, values_only=True):
        if isinstance(row[2], str):  
            fecha = datetime.strptime(row[2], '%Y-%m-%d')  # Convierte a datetime si es un string
        else:
            fecha = row[2]

        if str(row[0]).strip() == ci and fecha.month == mes and fecha.year == year:
            if isinstance(fecha, datetime):  # Verifica si 'fecha' es un datetime
                fecha = fecha.date()  # Solo extrae la fecha (sin hora)
            else:
                fecha = datetime.strptime(fecha, "%Y-%m-%d").date()

            hora_entrada = datetime.combine(fecha, row[3]) if isinstance(row[3], time) else row[3]
            hora_salida = datetime.combine(fecha, row[4]) if isinstance(row[4], time) else row[4]
            retrasos = row[5] if len(row) > 5 else "00:00:00"
            deduccion = float(row[6]) if len(row) > 6 and row[6] else 0.0
            
            for registro in registros:

                # Convertir registro[0] a datetime.date si es una cadena
                if isinstance(registro[0], str):
                    registro[0] = datetime.strptime(registro[0], "%Y-%m-%d").date()  # Convertir a datetime.date
                
                # Si 'fecha' es una cadena, convertirla a datetime.date
                if isinstance(fecha, str):
                    fecha = datetime.strptime(fecha, "%Y-%m-%d").date()

                if registro[0] == fecha:
                    horas_programadas = [datetime.strptime(hora_inicio, "%H:%M").time() for _, hora_inicio, _ in horario[registro[1]]]
                    retraso_minutos = sum([calcular_retraso(hora_entrada, datetime.combine(fecha, hora_programada)) for hora_programada in horas_programadas])
                    deduccion = calcular_deduccion(retraso_minutos)
                    if hora_salida is None:
                        horas_trabajadas = 0
                    else:
                        horas_trabajadas = round((hora_salida - hora_entrada).seconds / 3600, 2)                    
                    total_horas += horas_trabajadas
                    deducciones += deduccion
                    total_retrasos += timedelta(minutes=retraso_minutos)

                    registro[3] = horas_trabajadas
                    registro[4] = formatear_retraso(retraso_minutos)
                    registro[5] = deduccion
                    break

    # Ordenar los registros por fecha en orden ascendente
    registros.sort(key=lambda x: datetime.strptime(x[0], "%Y-%m-%d").date() if isinstance(x[0], str) else x[0])
    pago_por_hora = docentes[ci][1]
    total_ganado = round(total_horas * pago_por_hora, 2)
    neto_ganado = round(total_ganado - deducciones, 2)
    return registros, round(total_horas, 2), total_ganado, round(deducciones, 2), neto_ganado

def exportar_a_excel(ci, mes, year, plantilla_path, output_path):
    """Escribe los datos del reporte en una plantilla de Excel y guarda el archivo resultante."""
    # Cargar la plantilla de Excel
    wb = openpyxl.load_workbook(plantilla_path)
    ws = wb.active  # Puedes cambiar esto si la hoja tiene un nombre específico

    # Obtener los datos del reporte
    registros, total_horas, total_ganado, deducciones, neto_ganado = generar_reporte(ci, mes, year)

    # Datos del docente
    nombre_docente = docentes[ci][0]
    pago_por_hora = docentes[ci][1]

    # Obtener el horario del docente para los días de trabajo
    horario = obtener_horario(ci)
    dias_trabajo = [dia.capitalize() for dia in horario.keys()]
    dias_trabajo_str = " - ".join(dias_trabajo)

    # Escribir datos del docente en la plantilla
    ws["D6"].value = nombre_docente  # Nombre del docente
    ws["D8"].value = pago_por_hora  # Pago por hora
    ws["D10"].value = dias_trabajo_str  # Días que viene a trabajar el docente
    ws["J6"].value = ci  # CI del docente

    # Calcular el período de declaración
    fecha_inicio = f"01/{mes:02d}/{year}"
    fecha_fin = f"{(datetime(year, mes + 1, 1) - timedelta(days=1)).strftime('%d/%m/%Y')}"
    periodo_declaracion = f"{fecha_inicio} al {fecha_fin}"
    ws["J8"].value = periodo_declaracion  # Período de declaración

    # Obtener el estilo de las celdas de la fila 16 en las columnas B a H
    estilos_celdas_plantilla = [ws[f"{col}16"]._style for col in "BCDEFGH"]

    # Escribir los datos en las filas ya existentes de la hoja de Excel
    fila_inicio = 16  # Fila inicial para los datos de la tabla
    fila_fin = 31  # Fila final para los datos de la tabla

    for i, registro in enumerate(registros):
        fila = fila_inicio + i
        if fila > fila_fin:
            ws.insert_rows(fila)  # Insertar nueva fila si excede el límite

        # Aplicar estilo y valor a las columnas B a H
        for j, value in enumerate(registro):
            celda = ws.cell(row=fila, column=2 + j, value=value)
            celda._style = estilos_celdas_plantilla[j]

    # Llenar la tabla de deducciones desde la celda J17 y K17
    fila_deducciones_inicio = 17
    col_fecha_deduccion = 10  # Columna J
    col_monto_deduccion = 11  # Columna K
    fila_deducciones_actual = fila_deducciones_inicio

    for registro in registros:
        fecha = registro[0]
        monto_deduccion = registro[5]
        if monto_deduccion > 0:  # Solo incluir deducciones mayores a 0
            ws.cell(row=fila_deducciones_actual, column=col_fecha_deduccion, value=fecha)
            ws.cell(row=fila_deducciones_actual, column=col_monto_deduccion, value=monto_deduccion)
            fila_deducciones_actual += 1

    # Crear el directorio si no existe
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Guardar el archivo con los datos actualizados
    wb.save(output_path)
    print(f"Reporte guardado en: {output_path}")

def abrir_reporte(root_menu):
    """Abre la ventana para generar el reporte de horas trabajadas y deducciones."""
    verificar_archivo_asistencia()
    global docentes
    docentes = obtener_docentes()

    # Diccionario para mapear los nombres de los meses en español a números
    meses_espanol = {
        "Enero": 1,
        "Febrero": 2,
        "Marzo": 3,
        "Abril": 4,
        "Mayo": 5,
        "Junio": 6,
        "Julio": 7,
        "Agosto": 8,
        "Septiembre": 9,
        "Octubre": 10,
        "Noviembre": 11,
        "Diciembre": 12
    }

    # Evitar múltiples ventanas abiertas
    if hasattr(abrir_reporte, "ventana") and abrir_reporte.ventana.winfo_exists():
        abrir_reporte.ventana.lift()
        return

    def generar():
        """Genera el reporte y muestra los resultados en la tabla."""
        docente_nombre = combo_docente.get().strip()
        mes = combo_mes.get().strip()
        year = int(combo_year.get().strip())

        if not docente_nombre or not mes or not year:
            messagebox.showwarning("Datos incompletos", "Por favor, complete todos los campos antes de generar el reporte.")
            return

        ci = [ci for ci, info in docentes.items() if info[0] == docente_nombre][0]
        mes_numero = meses_espanol[mes]

        registros, total_horas, total_ganado, deducciones, neto_ganado = generar_reporte(ci, mes_numero, year)

        for item in tree.get_children():
            tree.delete(item)

        for registro in registros:
            if registro[4] != "00:00:00":
                tree.insert("", "end", values=registro, tags=('retraso',))
            else:
                tree.insert("", "end", values=registro)

        lbl_totales.config(text=f"Total Horas: {total_horas}\nTotal Ganado: Bs {total_ganado}\nDeducciones: Bs {deducciones}\nNeto Ganado: Bs {neto_ganado}")

        # Mostrar botones de exportación
        btn_exportar_excel.pack(side=tk.LEFT, padx=10)

    def exportar_excel():
        """Exporta el reporte a un archivo Excel."""
        docente_nombre = combo_docente.get().strip()
        mes = combo_mes.get().strip()
        year = int(combo_year.get().strip())
        ci = [ci for ci, info in docentes.items() if info[0] == docente_nombre][0]
        mes_numero = meses_espanol[mes]
        output_path = os.path.join(os.path.dirname(__file__), f'../reportes/reporte_{ci}_{mes}_{year}.xlsx')
        exportar_a_excel(ci, mes_numero, year, plantilla_excel, output_path)
        messagebox.showinfo("Exportar a Excel", f"Reporte guardado en: {output_path}")

    root_menu.withdraw()
    root = tk.Toplevel()
    abrir_reporte.ventana = root
    root.title("Generar Reporte")
    root.state('zoomed')
    root.configure(bg='#e0e0e0')
    
    style = ttk.Style()
    style.configure("Treeview.Heading", font=('Arial', 12, 'bold'), background='#d9d9d9', foreground='black')
    style.configure("Treeview", font=('Arial', 10), rowheight=25, background='#f0f0f0', foreground='black', fieldbackground='#f0f0f0')
    style.map('Treeview', background=[('selected', '#347083')], foreground=[('selected', 'white')])

    frame_main = tk.Frame(root, bg='#e0e0e0')
    frame_main.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

    # Filtros en una fila
    frame_filtros = tk.Frame(frame_main, bg='#e0e0e0')
    frame_filtros.pack(fill=tk.X, pady=10)

    tk.Label(frame_filtros, text="Docente:", bg='#e0e0e0', font=('Arial', 10)).pack(side=tk.LEFT, padx=5)
    combo_docente = ttk.Combobox(frame_filtros, values=[info[0] for info in docentes.values()], width=30)
    combo_docente.pack(side=tk.LEFT, padx=5)

    tk.Label(frame_filtros, text="Mes:", bg='#e0e0e0', font=('Arial', 10)).pack(side=tk.LEFT, padx=5)
    combo_mes = ttk.Combobox(frame_filtros, values=list(meses_espanol.keys()), width=15)
    combo_mes.pack(side=tk.LEFT, padx=5)

    tk.Label(frame_filtros, text="Año:", bg='#e0e0e0', font=('Arial', 10)).pack(side=tk.LEFT, padx=5)
    combo_year = ttk.Combobox(frame_filtros, values=obtener_anios_disponibles(), width=10)
    combo_year.pack(side=tk.LEFT, padx=5)

    tk.Button(frame_filtros, text="Generar Reporte", command=generar, bg='#212121', fg='white').pack(side=tk.LEFT, padx=10)

    # Frame para la tabla y los resultados
    frame_resultados = tk.Frame(frame_main, bg='#e0e0e0')
    frame_resultados.pack(fill=tk.BOTH, expand=True)

    frame_table = tk.Frame(frame_resultados, bg='#e0e0e0')
    frame_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)

    frame_totales = tk.Frame(frame_resultados, bg='#e0e0e0')
    frame_totales.pack(side=tk.BOTTOM, fill=tk.X, padx=10)

    # Tabla para mostrar el reporte
    columns = ["Fecha", "Día", "Materias", "Horas Trabajadas", "Retrasos", "Descuento BS", "Modalidad"]
    tree = ttk.Treeview(frame_table, columns=columns, show="headings", height=15)
    
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=130, anchor="center")
    
    tree.pack(pady=10, fill=tk.BOTH, expand=True)

    # Cambiar el color de la columna de retrasos a rojo
    style.configure("Treeview", foreground="black", font=('Arial', 10))
    style.map("Treeview", foreground=[('!selected', 'black'), ('selected', 'white')])
    tree.tag_configure('retraso', foreground='red')

    lbl_totales = tk.Label(frame_totales, text="", bg='#e0e0e0', font=('Arial', 12, 'bold'), justify=tk.LEFT)
    lbl_totales.pack(pady=10, anchor="w")

    # Botón para exportar a Excel
    btn_exportar_excel = tk.Button(frame_filtros, text="Exportar a Excel", command=exportar_excel, bg='#212121', fg='white')
    btn_exportar_excel.pack_forget()  # Ocultar inicialmente

    # Botón para volver al menú
    tk.Button(root, text="Volver al Menú", command=lambda: volver_al_menu(root, root_menu), 
              bg='#212121', fg='white', width=20).pack(pady=10)

    root.protocol("WM_DELETE_WINDOW", lambda: volver_al_menu(root, root_menu))  # Manejar cierre de ventana
    root.mainloop()

def volver_al_menu(root_actual, root_menu):
    """Cierra la ventana actual y muestra nuevamente el menú principal."""
    root_actual.destroy()
    root_menu.deiconify()
    root_menu.state('zoomed')  # Maximiza la ventana del menú principal al volver
