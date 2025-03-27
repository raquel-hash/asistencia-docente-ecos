import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
import os

archivo_materias = os.path.join(os.path.dirname(__file__), '../data/materias.xlsx')

def verificar_archivo_materias():
    """Verifica si el archivo de materias existe, si no lo crea vacío con encabezados."""
    if not os.path.exists(archivo_materias):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Materia"])
        wb.save(archivo_materias)

def actualizar_lista_materias(tree):
    """Actualiza la tabla de materias en la ventana."""
    for row in tree.get_children():
        tree.delete(row)

    wb = openpyxl.load_workbook(archivo_materias)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        tree.insert("", "end", values=row)

def abrir_lista_materias(root_menu):
    """Abre la ventana con la lista de materias y el formulario de registro."""
    verificar_archivo_materias()  # Asegura que el archivo existe antes de abrir la ventana
    
    # Evitar múltiples ventanas abiertas
    if hasattr(abrir_lista_materias, "ventana") and abrir_lista_materias.ventana.winfo_exists():
        abrir_lista_materias.ventana.lift()  # Lleva la ventana al frente si ya existe
        return

    root_menu.withdraw()  # Oculta el menú principal
    root = tk.Toplevel()
    abrir_lista_materias.ventana = root  # Guarda la referencia a la ventana para evitar duplicados
    root.title("Lista de Materias")
    root.geometry("800x600")
    root.state('zoomed')  # Ocupa toda la pantalla
    root.configure(bg='#e0e0e0')
    
    style = ttk.Style()
    style.configure("Treeview.Heading", font=('Arial', 12, 'bold'), background='#d9d9d9', foreground='black')
    style.configure("Treeview", font=('Arial', 10), rowheight=25, background='#f0f0f0', foreground='black', fieldbackground='#f0f0f0')

    # Frame principal
    frame_main = tk.Frame(root, bg='#e0e0e0')
    frame_main.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

    frame_form = tk.Frame(frame_main, bg='#e0e0e0')
    frame_form.pack(side=tk.LEFT, fill=tk.Y, padx=20, pady=20)

    frame_table = tk.Frame(frame_main, bg='#e0e0e0')
    frame_table.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=20, pady=20)

    # Título
    # tk.Label(frame_main, text="Gestión de Materias", font=("Arial", 24), bg='#e0e0e0').pack(pady=20)

    # Formulario de registro de materias
    tk.Label(frame_form, text="Materia:", bg='#e0e0e0', font=('Arial', 10)).grid(row=0, column=0, padx=5, pady=5, sticky="e")
    entry_materia = tk.Entry(frame_form)
    entry_materia.grid(row=0, column=1, padx=5, pady=5)

    def agregar_materia():
        """Agrega una nueva materia al archivo y actualiza la lista."""
        materia = entry_materia.get()

        if not materia:
            messagebox.showwarning("Campo incompleto", "El campo de materia es obligatorio.")
            return

        wb = openpyxl.load_workbook(archivo_materias)
        ws = wb.active
        ws.append([materia])
        wb.save(archivo_materias)

        actualizar_lista_materias(tree)
        entry_materia.delete(0, tk.END)

    def eliminar_materia():
        """Elimina la materia seleccionada de la tabla y del archivo."""
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("Seleccionar materia", "Debe seleccionar una materia para eliminar.")
            return
        
        materia = tree.item(selected_item, "values")[0]

        wb = openpyxl.load_workbook(archivo_materias)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if row[0].value == materia:
                ws.delete_rows(row[0].row)
                break
        wb.save(archivo_materias)

        actualizar_lista_materias(tree)

    tk.Button(frame_form, text="Agregar Materia", command=agregar_materia, bg='#212121', fg='white').grid(row=1, column=0, pady=10)
    tk.Button(frame_form, text="Eliminar Materia", command=eliminar_materia, bg='#212121', fg='white').grid(row=1, column=1, pady=10)

    # Tabla de materias
    columns = ["Materia"]
    tree = ttk.Treeview(frame_table, columns=columns, show="headings", height=10)
    
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=200, anchor="center")
    
    tree.pack(pady=10, fill=tk.BOTH, expand=True)

    actualizar_lista_materias(tree)

    # Botón para volver al menú
    tk.Button(root, text="Volver al Menú", command=lambda: volver_al_menu(root, root_menu), 
              bg='#212121', fg='white', width=20).pack(pady=10)

    root.protocol("WM_DELETE_WINDOW", lambda: volver_al_menu(root, root_menu))  # Manejar cierre de ventana
    root.mainloop()

def volver_al_menu(root_actual, root_menu):
    """Cierra la ventana actual y muestra nuevamente el menú principal."""
    root_actual.destroy()
    root_menu.deiconify()
    root_menu.state('zoomed')  # Maximiza la ventana del menú principal al volver
